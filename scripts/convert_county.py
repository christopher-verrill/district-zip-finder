"""
Convert county_zip_list_csv.xlsx -> public/data/county.json

Output shape matches the other district types so App.jsx can use it without changes:
  {
    "STATE/COUNTY NAME": [
      { "zip": "27284", "zip_pop": 12345, "district_pop": 10000, "overlap": 0.81 },
      ...
    ],
    ...
  }

Source columns (header row): COUNTY, zip, zip pop, COUNTY POP IN ZIP, overlap

Notes:
- The source file is named .csv but is actually an .xlsx. Rename it or point
  INPUT_PATH at the .xlsx version.
- 31 rows have overlap '#DIV/0!' because zip_pop is 0 (mostly water-only ZIPs in
  CT's planning regions). These are dropped.
- "COUNTY POP IN ZIP" maps to "district_pop" in the output so the app code that
  expects that key for federal/state district data works unchanged.
"""

import json
import os
import sys
from openpyxl import load_workbook

# --- Paths ---
# Adjust if your source file lives elsewhere.
INPUT_PATH = os.path.join(
    os.path.dirname(__file__), "source", "county_zip_list_csv.xlsx"
)
OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "public", "data", "county.json"
)


def main():
    if not os.path.exists(INPUT_PATH):
        print(f"ERROR: Cannot find input file: {INPUT_PATH}", file=sys.stderr)
        print("Place the source file at scripts/source/county_zip_list_csv.xlsx", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {INPUT_PATH}...")
    wb = load_workbook(INPUT_PATH, read_only=True)
    ws = wb.active

    out = {}
    rows_processed = 0
    rows_dropped = 0
    dropped_examples = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Take only the first 5 columns in case the source has trailing junk
        if len(row) < 5:
            rows_dropped += 1
            continue
        county, zip_code, zip_pop, county_pop_in_zip, overlap = row[:5]

        # Skip blank/trailing rows
        if county is None or zip_code is None:
            rows_dropped += 1
            continue

        # Drop rows with non-numeric overlap (the #DIV/0! cases)
        if not isinstance(overlap, (int, float)):
            rows_dropped += 1
            if len(dropped_examples) < 5:
                dropped_examples.append((county, zip_code, overlap))
            continue

        # Defensive: also drop zip_pop=0 rows since they can't contribute to coverage
        if not zip_pop:
            rows_dropped += 1
            continue

        # Pad zip to 5 digits as a string (preserves leading zeros for NE states)
        zip_str = str(int(zip_code)).zfill(5)

        entry = {
            "zip": zip_str,
            "zip_pop": int(zip_pop),
            "district_pop": int(county_pop_in_zip) if county_pop_in_zip is not None else 0,
            "overlap": float(overlap),
        }

        if county not in out:
            out[county] = []
        out[county].append(entry)
        rows_processed += 1

    # Sort each county's ZIPs by overlap desc (matches how the app sorts them)
    for county in out:
        out[county].sort(key=lambda z: z["overlap"], reverse=True)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"Wrote {OUTPUT_PATH} ({size_kb:.1f} KB)")
    print(f"  Counties: {len(out)}")
    print(f"  Rows kept: {rows_processed}")
    print(f"  Rows dropped: {rows_dropped}")
    if dropped_examples:
        print("  Dropped examples:")
        for ex in dropped_examples:
            print(f"    {ex}")


if __name__ == "__main__":
    main()