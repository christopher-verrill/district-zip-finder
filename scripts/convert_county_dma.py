"""
Convert county_and_zip_crosswalk.xlsx -> public/data/county_dma_crosswalk.json

Output shape matches the other DMA crosswalk files so App.jsx works unchanged:
  {
    "STATE/COUNTY NAME": [
      { "dma_name": "Los Angeles CA", "pop_in_cd": 7232495, "total_dma_pop": 13436893 },
      ...
    ],
    ...
  }

Source columns: COUNTY, DMA, dma pop, County pop in DMA, (overlap formula)

Notes:
- The source uses XLOOKUP formulas for `dma pop`. Excel caches the evaluated
  values when the file is saved, and openpyxl with data_only=True reads those
  cached values. If a fresh export from the source spreadsheet ever shows
  `None` in column C, open the file in Excel and re-save before running this.
- 16 rows with null county or DMA are dropped (likely trailing rows).
- 54 crosswalk entries for Guam and Puerto Rico are kept in the output but
  will be ignored by the app since they don't have matching county.json keys.
"""

import json
import os
import sys
from openpyxl import load_workbook

INPUT_PATH = os.path.join(
    os.path.dirname(__file__), "source", "county_and_zip_crosswalk.xlsx"
)
OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "public", "data", "county_dma_crosswalk.json"
)


def main():
    if not os.path.exists(INPUT_PATH):
        print(f"ERROR: Cannot find input file: {INPUT_PATH}", file=sys.stderr)
        print(
            "Place the source file at scripts/source/county_and_zip_crosswalk.xlsx",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Reading {INPUT_PATH}...")
    # data_only=True reads the cached results of Excel formulas (XLOOKUP).
    wb = load_workbook(INPUT_PATH, read_only=True, data_only=True)
    ws = wb.active

    out = {}
    rows_kept = 0
    rows_dropped = 0
    dropped_examples = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        county = row[0]
        dma = row[1]
        dma_pop = row[2]
        pop_in_county = row[3]

        # Drop trailing or empty rows
        if not county or not dma:
            rows_dropped += 1
            continue

        # Drop TargetSmart's "Not Available" bucket — it's records with no DMA
        # assignment, not a real DMA. It appears in 3,100+ counties and would
        # corrupt the waste math and break the boundary-map matching.
        if dma == "Not Available":
            rows_dropped += 1
            continue

        # Drop rows with missing/non-numeric values (would happen if formulas
        # weren't cached). Capture examples to help debug.
        if not isinstance(dma_pop, (int, float)) or not isinstance(
            pop_in_county, (int, float)
        ):
            rows_dropped += 1
            if len(dropped_examples) < 3:
                dropped_examples.append((county, dma, dma_pop, pop_in_county))
            continue

        entry = {
            "dma_name": dma,
            "pop_in_cd": int(pop_in_county),
            "total_dma_pop": int(dma_pop),
        }
        out.setdefault(county, []).append(entry)
        rows_kept += 1

    # Sort each county's DMAs by pop_in_cd descending (largest overlap first)
    for county in out:
        out[county].sort(key=lambda d: d["pop_in_cd"], reverse=True)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"Wrote {OUTPUT_PATH} ({size_kb:.1f} KB)")
    print(f"  Counties: {len(out)}")
    print(f"  Rows kept: {rows_kept}")
    print(f"  Rows dropped: {rows_dropped}")
    if dropped_examples:
        print("  Dropped (non-numeric) examples:")
        for ex in dropped_examples:
            print(f"    {ex}")
        print(
            "  If you see these, the source spreadsheet has uncached formulas."
        )
        print("  Open it in Excel and save, then re-run this script.")


if __name__ == "__main__":
    main()
