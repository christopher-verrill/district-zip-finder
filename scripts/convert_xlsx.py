#!/usr/bin/env python3
import json, os, sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

INPUT_PATH = os.path.join(os.path.dirname(__file__), "source", "state_house.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "public", "data", "state_house.json")

def convert(input_path, output_path):
    print(f"Reading {input_path} ...")
    wb = load_workbook(input_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Read header row and find columns by name flexibly
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    print("Columns found:", rows[0])

    def col(names):
        for name in names:
            for i, h in enumerate(header):
                if name.lower() in h:
                    return i
        return None

    idx_cd       = col(["house district", "cd path", "district"])
    idx_zip      = col(["zip"])
    idx_zip_pop  = col(["zip pop"])
    idx_dist_pop = col(["hd pop", "pop of congressional", "district pop"])
    idx_overlap  = col(["overlap"])

    print(f"Column indices — district:{idx_cd}, zip:{idx_zip}, zip_pop:{idx_zip_pop}, dist_pop:{idx_dist_pop}, overlap:{idx_overlap}")

    data = {}
    skipped = 0

    for row in rows[1:]:
        try:
            cd          = row[idx_cd]
            zip_code    = row[idx_zip]
            zip_pop     = row[idx_zip_pop]
            district_pop = row[idx_dist_pop]
            overlap     = row[idx_overlap]
        except IndexError:
            skipped += 1
            continue

        if not cd or not zip_code or not isinstance(overlap, (int, float)):
            skipped += 1
            continue

        if cd not in data:
            data[cd] = []
        data[cd].append({
            "zip": str(zip_code).zfill(5),
            "zip_pop": zip_pop if isinstance(zip_pop, (int, float)) else 0,
            "district_pop": district_pop if isinstance(district_pop, (int, float)) else 0,
            "overlap": round(float(overlap), 6),
        })

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(data, f, separators=(",", ":"))

    size_kb = os.path.getsize(output_path) / 1024
    print(f"Done. {len(data)} districts written to {output_path} ({size_kb:.0f} KB). Skipped {skipped} rows.")

if __name__ == "__main__":
    if not os.path.exists(INPUT_PATH):
        print(f"Input file not found: {INPUT_PATH}")
        print("Place your XLSX at scripts/source/state_house.xlsx and re-run.")
        sys.exit(1)
    convert(INPUT_PATH, OUTPUT_PATH)